import vk_api
import json
import re
import logging
import sqlite3
import time
import asyncio
import urllib.parse
import sys
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, CallbackContext
from telegram.ext import JobQueue
from telegram.error import TelegramError, NetworkError
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import io
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Загружаем переменные окружения из .env файла
load_dotenv()

# ---------------- Настройка логирования ----------------
# Отключаем все логи кроме ошибок
logging.getLogger("httpx").setLevel(logging.ERROR)
logging.getLogger("telegram").setLevel(logging.ERROR)
logging.getLogger("apscheduler").setLevel(logging.ERROR)
logging.getLogger("vk_api").setLevel(logging.ERROR)

# Минимальное логирование для нашего бота
logging.basicConfig(
    format='%(asctime)s - %(message)s',
    level=logging.INFO,
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ---------------- Токены ----------------
# Токены загружаются из переменных окружения (.env файл)
VK_TOKEN = os.getenv("VK_TOKEN", "")
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN", "")

# Проверка наличия токенов
if not VK_TOKEN:
    logger.error("❌ VK_TOKEN не найден в переменных окружения! Укажите его в файле .env")
if not TELEGRAM_TOKEN:
    logger.error("❌ TELEGRAM_TOKEN не найден в переменных окружения! Укажите его в файле .env")

# Глобальная переменная для отслеживания состояния выполнения
is_checking = False
bot_start_time = None

# ---------------- Excel файлы ----------------
POSTS_EXCEL_FILE = "checked_posts.xlsx"
COMMENTS_EXCEL_FILE = "found_comments.xlsx"


# ---------------- Функции для работы с Excel ----------------
def format_excel_file(file_path, sheet_name="Sheet1"):
    """Форматирует Excel файл: настраивает ширину колонок, заголовки и т.д."""
    try:
        if not os.path.exists(file_path):
            return False

        # Загружаем workbook
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        # Настраиваем ширину колонок в зависимости от файла
        if file_path == POSTS_EXCEL_FILE:
            # Форматирование для файла постов
            column_widths = {
                'A': 35,  # Ссылка на группу
                'B': 35,  # Ссылка на пост
                'C': 50,  # Текст поста
                'D': 20,  # Дата проверки
            }
        else:
            # Форматирование для файла комментариев
            column_widths = {
                'A': 25,  # Имя пользователя
                'B': 35,  # Ссылка на пользователя
                'C': 20,  # Город
                'D': 50,  # Текст комментария
                'E': 35,  # Ссылка на комментарий
                'F': 20,  # Ключевое слово
                'G': 20,  # Дата обнаружения
            }

        # Применяем ширину колонок
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Форматируем заголовки
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Включаем фильтры для заголовков
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

        # Замораживаем первую строку (заголовки)
        ws.freeze_panes = 'A2'

        # Сохраняем изменения
        wb.save(file_path)
        logger.info(f"✅ Отформатирован Excel файл: {file_path}")
        return True

    except Exception as e:
        logger.error(f"❌ Ошибка форматирования Excel файла {file_path}: {e}")
        return False


def init_excel_files():
    """Инициализирует Excel файлы с заголовками и форматированием"""
    # Файл для проверенных постов
    if not os.path.exists(POSTS_EXCEL_FILE):
        df_posts = pd.DataFrame(columns=[
            'Ссылка на группу',
            'Ссылка на пост',
            'Текст поста (первые 50 символов)',
            'Дата проверки'
        ])
        df_posts.to_excel(POSTS_EXCEL_FILE, index=False, engine='openpyxl')
        format_excel_file(POSTS_EXCEL_FILE)
        logger.info("✅ Создан Excel файл для проверенных постов")

    # Файл для найденных комментариев
    if not os.path.exists(COMMENTS_EXCEL_FILE):
        df_comments = pd.DataFrame(columns=[
            'Имя пользователя',
            'Ссылка на страницу пользователя',
            'Город',
            'Текст комментария',
            'Ссылка на комментарий',
            'Найденное ключевое слово',
            'Дата обнаружения'
        ])
        df_comments.to_excel(COMMENTS_EXCEL_FILE, index=False, engine='openpyxl')
        format_excel_file(COMMENTS_EXCEL_FILE)
        logger.info("✅ Создан Excel файл для найденных комментариев")


def add_post_to_excel(group_domain, group_id, post_id, post_text):
    """Добавляет проверенный пост в Excel файл"""
    try:
        # Читаем существующий файл
        df = pd.read_excel(POSTS_EXCEL_FILE, engine='openpyxl')

        # Проверяем, нет ли уже этого поста
        post_exists = ((df['Ссылка на пост'] == f"https://vk.com/wall-{group_id}_{post_id}")).any()

        if not post_exists:
            # Формируем данные для нового поста
            group_link = f"https://vk.com/{group_domain}"
            post_link = f"https://vk.com/wall-{group_id}_{post_id}"
            post_preview = post_text[:50] + "..." if len(post_text) > 50 else post_text
            check_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            new_post = {
                'Ссылка на группу': group_link,
                'Ссылка на пост': post_link,
                'Текст поста (первые 50 символов)': post_preview,
                'Дата проверки': check_date
            }

            # Добавляем новую строку
            df = pd.concat([df, pd.DataFrame([new_post])], ignore_index=True)

            # Сохраняем обратно в Excel
            df.to_excel(POSTS_EXCEL_FILE, index=False, engine='openpyxl')

            # Форматируем файл после добавления данных
            format_excel_file(POSTS_EXCEL_FILE)

            logger.info(f"✅ Добавлен пост в Excel: {group_domain} - {post_id}")
            return True
        return False

    except Exception as e:
        logger.error(f"❌ Ошибка при добавлении поста в Excel: {e}")
        return False


def add_comment_to_excel(comment_data):
    """Добавляет найденный комментарий в Excel файл"""
    try:
        # Читаем существующий файл
        df = pd.read_excel(COMMENTS_EXCEL_FILE, engine='openpyxl')

        # Проверяем, нет ли уже этого комментария
        comment_exists = ((df['Ссылка на комментарий'] == comment_data['comment_link'])).any()

        if not comment_exists:
            # Подготавливаем данные для Excel в новом порядке
            excel_data = {
                'Имя пользователя': comment_data['user_name'],
                'Ссылка на страницу пользователя': comment_data['user_link'],
                'Город': comment_data['city'],
                'Текст комментария': comment_data['text'],
                'Ссылка на комментарий': comment_data['comment_link'],
                'Найденное ключевое слово': comment_data['keyword'],
                'Дата обнаружения': comment_data['detection_date']
            }

            # Добавляем новую строку
            df = pd.concat([df, pd.DataFrame([excel_data])], ignore_index=True)

            # Сохраняем обратно в Excel
            df.to_excel(COMMENTS_EXCEL_FILE, index=False, engine='openpyxl')

            # Форматируем файл после добавления данных
            format_excel_file(COMMENTS_EXCEL_FILE)

            logger.info(f"✅ Добавлен комментарий в Excel: {comment_data['user_name']}")
            return True
        return False

    except Exception as e:
        logger.error(f"❌ Ошибка при добавлении комментария в Excel: {e}")
        return False


def get_excel_stats():
    """Возвращает статистику по Excel файлам"""
    try:
        posts_count = 0
        comments_count = 0

        if os.path.exists(POSTS_EXCEL_FILE):
            df_posts = pd.read_excel(POSTS_EXCEL_FILE, engine='openpyxl')
            posts_count = len(df_posts)

        if os.path.exists(COMMENTS_EXCEL_FILE):
            df_comments = pd.read_excel(COMMENTS_EXCEL_FILE, engine='openpyxl')
            comments_count = len(df_comments)

        return posts_count, comments_count
    except Exception as e:
        logger.error(f"❌ Ошибка получения статистики Excel: {e}")
        return 0, 0


# ---------------- Улучшенная настройка VK API с повторными попытками ----------------
def create_vk_session_with_retry():
    """Создает VK сессию с настройками для повторных попыток"""
    session = vk_api.VkApi(
        token=VK_TOKEN,
        api_version='5.131'
    )

    # Настройка повторных попыток для requests
    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.http.mount("http://", adapter)
    session.http.mount("https://", adapter)

    # Увеличиваем таймауты
    session.http.timeout = 30

    return session


# ---------------- База данных ----------------
def init_db():
    conn = sqlite3.connect('vk_monitor.db')
    cursor = conn.cursor()

    # Таблица для групп ВК
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS vk_groups (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        domain TEXT UNIQUE NOT NULL,
        group_id INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')

    # Проверяем наличие столбца group_id и добавляем его, если отсутствует
    cursor.execute("PRAGMA table_info(vk_groups)")
    columns = [column[1] for column in cursor.fetchall()]
    if 'group_id' not in columns:
        cursor.execute('ALTER TABLE vk_groups ADD COLUMN group_id INTEGER')

    # Таблица для ключевых слов
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS keywords (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        keyword TEXT UNIQUE NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')

    # Таблица для чатов Telegram (как личные, так и группы)
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS telegram_chats (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chat_id INTEGER UNIQUE NOT NULL,
        chat_type TEXT NOT NULL,
        chat_title TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')

    # Таблица для статистики
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS bot_stats (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        total_comments INTEGER DEFAULT 0,
        last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')

    # Инициализируем статистику, если нет записей
    cursor.execute('SELECT COUNT(*) FROM bot_stats')
    if cursor.fetchone()[0] == 0:
        cursor.execute('INSERT INTO bot_stats (total_comments) VALUES (0)')

    conn.commit()
    conn.close()


def get_db_connection():
    return sqlite3.connect('vk_monitor.db')


# ---------------- Функции для работы со статистикой ----------------
def get_total_comments_count():
    """Получает общее количество найденных комментариев"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT total_comments FROM bot_stats WHERE id = 1')
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else 0


def update_total_comments_count(count):
    """Обновляет общее количество найденных комментариев"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('UPDATE bot_stats SET total_comments = ?, last_updated = CURRENT_TIMESTAMP WHERE id = 1', (count,))
    conn.commit()
    conn.close()


def increment_total_comments_count():
    """Увеличивает счетчик найденных комментариев на 1"""
    current_count = get_total_comments_count()
    update_total_comments_count(current_count + 1)


# ---------------- Инициализация VK API ----------------
try:
    vk_session = create_vk_session_with_retry()
    vk = vk_session.get_api()
    print("✓ VK API подключен")
except Exception as e:
    print(f"✗ Ошибка VK API: {e}")
    vk = None


# ---------------- Функция для получения статуса бота ----------------
def get_bot_status():
    """Возвращает статус бота"""
    global bot_start_time
    status = "🟢 ОНЛАЙН"

    if bot_start_time:
        uptime = datetime.now() - bot_start_time
        hours, remainder = divmod(uptime.total_seconds(), 3600)
        minutes, seconds = divmod(remainder, 60)
        uptime_str = f"{int(hours)}ч {int(minutes)}м {int(seconds)}с"
    else:
        uptime_str = "неизвестно"

    groups = get_groups()
    keywords = get_keywords()
    chats = get_all_chats()
    total_comments = get_total_comments_count()

    # Статистика из Excel файлов
    excel_posts, excel_comments = get_excel_stats()

    status_info = (
        f"{status}\n"
        f"⏰ Время работы: {uptime_str}\n"
        f"📊 Групп ВК: {len(groups)}\n"
        f"🔍 Ключевых слов: {len(keywords)}\n"
        f"💬 Чатов для уведомлений: {len(chats)}\n"
        f"📈 Всего найдено комментариев: {total_comments}\n"
        f"📁 Постов в Excel: {excel_posts}\n"
        f"📁 Комментариев в Excel: {excel_comments}\n"
        f"🕒 Последняя проверка: {datetime.now().strftime('%H:%M:%S')}"
    )

    return status_info


# ---------------- УПРОЩЕННАЯ ПРОВЕРКА ДОСТУПА - РАЗРЕШАЕМ ВСЕМ ----------------
async def check_access(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Упрощенная проверка доступа - разрешаем всем пользователям"""
    return True


# ---------------- Функция для получения аватарки пользователя ----------------
def get_user_photo_url(user_id):
    """Получает URL аватарки пользователя VK"""
    try:
        user_info = vk.users.get(
            user_ids=user_id,
            fields="photo_100,photo_200,photo_max"
        )
        if user_info:
            user = user_info[0]
            # Пробуем получить фото в порядке приоритета: photo_200, photo_100, photo_max
            photo_url = user.get('photo_200') or user.get('photo_100') or user.get('photo_max')
            return photo_url
    except Exception as e:
        return None


# ---------------- Функция для загрузки изображения ----------------
async def download_photo(url):
    """Загружает изображение по URL"""
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            return io.BytesIO(response.content)
    except Exception as e:
        return None


# ---------------- Функция для извлечения идентификатора группы из URL ----------------
def extract_group_id_from_url(url):
    """
    Извлекает идентификатор группы из URL ВКонтакте.
    Поддерживает форматы:
    - https://vk.com/group_name
    - https://vk.com/club123
    - https://vk.com/public123
    - vk.com/group_name
    - @group_name
    - group_name
    """
    # Если это короткое имя (начинается с @)
    if url.startswith('@'):
        return url[1:]

    # Если это полный URL
    if 'vk.com/' in url:
        # Извлекаем часть после vk.com/
        path = url.split('vk.com/')[1]
        # Удаляем параметры запроса, если есть
        path = path.split('?')[0]
        # Удаляем слеши в начале и конце
        path = path.strip('/')
        return path

    # Если это просто имя группы (без URL)
    return url


# ---------------- Функция для проверки ключевых слов ----------------
def contains_keyword(text, keywords):
    """
    Проверяет, содержит ли текст любое из ключевых слов.
    Учитывает разные регистры и исключает случаи, когда ключевое слово является частью другого слова.
    """
    if not text or not keywords:
        return False, None

    for keyword in keywords:
        # Используем регулярное выражение для поиска целых слов с игнорированием регистра
        pattern = r'\b' + re.escape(keyword) + r'\b'
        if re.search(pattern, text, re.IGNORECASE):
            return True, keyword

    return False, None


# ---------------- Клавиатура ----------------
def get_main_keyboard():
    keyboard = [
        [KeyboardButton("Добавить группу"), KeyboardButton("Добавить ключевое слово")],
        [KeyboardButton("Список групп"), KeyboardButton("Список ключевых слов")],
        [KeyboardButton("Проверить сейчас"), KeyboardButton("Удалить группу"),
         KeyboardButton("Удалить ключевое слово")],
        [KeyboardButton("Удалить все ключевые слова"), KeyboardButton("Статус"), KeyboardButton("Экспорт в Excel")],
        [KeyboardButton("Добавить чат"), KeyboardButton("Удалить чат"), KeyboardButton("Список чатов")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, is_persistent=True)


def get_admin_keyboard():
    keyboard = [
        [KeyboardButton("Статус"), KeyboardButton("Экспорт в Excel")],
        [KeyboardButton("Добавить группу"), KeyboardButton("Добавить ключевое слово")],
        [KeyboardButton("Список групп"), KeyboardButton("Список ключевых слов")],
        [KeyboardButton("Проверить сейчас"), KeyboardButton("Удалить группу"),
         KeyboardButton("Удалить ключевое слово")],
        [KeyboardButton("Удалить все ключевые слова")],
        [KeyboardButton("Добавить чат"), KeyboardButton("Удалить чат"), KeyboardButton("Список чатов")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, is_persistent=True)


# ---------------- Команды ----------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Всегда разрешаем доступ
    user = update.effective_user
    chat_type = update.effective_chat.type

    if chat_type in ['group', 'supergroup']:
        # Бот добавлен в группу
        chat_id = update.effective_chat.id
        chat_title = update.effective_chat.title

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'INSERT OR IGNORE INTO telegram_chats (chat_id, chat_type, chat_title) VALUES (?, ?, ?)',
            (chat_id, chat_type, chat_title)
        )
        conn.commit()
        conn.close()

        await update.message.reply_html(
            f"👋 Приветствую участников группы {chat_title}!\n\n"
            "Я бот для мониторинга комментариев ВКонтакте. "
            "Теперь эта группа будет получать уведомления о найденных комментариях.\n\n"
            f"{get_bot_status()}\n\n"
            "Для управления настройками используйте кнопки ниже:",
            reply_markup=get_admin_keyboard()
        )
    else:
        # Личный чат - разрешаем всем
        await update.message.reply_html(
            f"Привет, {user.mention_html()}!\n\n"
            "Я бот для мониторинга комментариев ВКонтакте.\n"
            "Я проверяю последние 20 постов в указанных группах на наличие ключевых слов.\n\n"
            f"{get_bot_status()}\n\n"
            "Используй кнопки ниже для управления мной:",
            reply_markup=get_main_keyboard()
        )


async def keyboard_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает клавиатуру снова"""
    # Всегда разрешаем доступ
    chat_type = update.effective_chat.type
    if chat_type in ['group', 'supergroup']:
        await update.message.reply_text("Клавиатура активирована", reply_markup=get_admin_keyboard())
    else:
        await update.message.reply_text("Клавиатура активирована", reply_markup=get_main_keyboard())


# ---------------- Утилиты базы данных ----------------
def add_chat_to_db(chat_id: int, chat_type: str, chat_title: str = None):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        'INSERT OR IGNORE INTO telegram_chats (chat_id, chat_type, chat_title) VALUES (?, ?, ?)',
        (chat_id, chat_type, chat_title)
    )
    conn.commit()
    conn.close()


def remove_chat_from_db(chat_id: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM telegram_chats WHERE chat_id = ?', (chat_id,))
    conn.commit()
    conn.close()


def is_chat_in_db(chat_id: int):
    """Проверяет, есть ли чат в базе данных"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM telegram_chats WHERE chat_id = ?', (chat_id,))
    result = cursor.fetchone()
    conn.close()
    return result is not None


def get_all_chats():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT chat_id, chat_type, chat_title FROM telegram_chats')
    chats = cursor.fetchall()
    conn.close()
    return chats


def get_chats_list_text():
    """Возвращает форматированный список чатов"""
    chats = get_all_chats()
    if not chats:
        return "📭 Список чатов для уведомлений пуст."

    chat_list = []
    for i, (chat_id, chat_type, chat_title) in enumerate(chats, 1):
        chat_type_emoji = "👥" if chat_type in ['group', 'supergroup'] else "👤"
        chat_name = chat_title if chat_title else f"Личный чат (ID: {chat_id})"
        chat_list.append(f"{i}. {chat_type_emoji} {chat_name} (ID: {chat_id})")

    return "📋 Чаты для уведомлений:\n" + "\n".join(chat_list)


def get_groups():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT domain, group_id FROM vk_groups')
    groups = [(row[0], row[1]) for row in cursor.fetchall()]
    conn.close()
    return groups


def get_keywords():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT keyword FROM keywords')
    keywords = [row[0] for row in cursor.fetchall()]
    conn.close()
    return keywords


def add_group(domain: str, group_id: int = None):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('INSERT OR IGNORE INTO vk_groups (domain, group_id) VALUES (?, ?)', (domain, group_id))
    conn.commit()
    conn.close()


def add_keyword(keyword: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('INSERT OR IGNORE INTO keywords (keyword) VALUES (?)', (keyword,))
    conn.commit()
    conn.close()


def delete_group(domain: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM vk_groups WHERE domain = ?', (domain,))
    conn.commit()
    conn.close()


def delete_keyword(keyword: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM keywords WHERE keyword = ?', (keyword,))
    conn.commit()
    conn.close()


def delete_all_keywords():
    """Удаляет все ключевые слова из базы данных"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM keywords')
    conn.commit()
    conn.close()
    logger.info("✅ Все ключевые слова удалены из базы данных")


# ---------------- Улучшенная функция безопасного VK запроса ----------------
async def safe_vk_request(func, *args, **kwargs):
    """Безопасный вызов VK API с обработкой ошибок"""
    max_retries = 3
    retry_delay = 2

    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)
        except vk_api.exceptions.ApiError as e:
            if attempt == max_retries - 1:
                raise
            await asyncio.sleep(retry_delay * (attempt + 1))
        except (requests.exceptions.RequestException, ConnectionError, TimeoutError) as e:
            if attempt == max_retries - 1:
                raise
            await asyncio.sleep(retry_delay * (attempt + 1))
        except Exception as e:
            raise

    return None


# ---------------- Улучшенная функция отправки уведомлений с фото ----------------
async def send_notification_with_photo(context: CallbackContext, text_message: str, photo_url: str = None):
    """Улучшенная функция отправки уведомлений с фото пользователя под текстом"""
    chats = get_all_chats()

    for chat_id, chat_type, chat_title in chats:
        max_retries = 3
        for attempt in range(max_retries):
            try:
                # Если есть фото, отправляем его с текстом как caption
                if photo_url:
                    photo_data = await download_photo(photo_url)
                    if photo_data:
                        await context.bot.send_photo(
                            chat_id=chat_id,
                            photo=photo_data,
                            caption=text_message,
                            parse_mode='HTML'
                        )
                    else:
                        # Если не удалось загрузить фото, отправляем только текст
                        await context.bot.send_message(
                            chat_id=chat_id,
                            text=text_message,
                            disable_web_page_preview=True,
                            parse_mode='HTML'
                        )
                else:
                    # Если нет фото, отправляем только текст
                    await context.bot.send_message(
                        chat_id=chat_id,
                        text=text_message,
                        disable_web_page_preview=True,
                        parse_mode='HTML'
                    )
                break
            except NetworkError as e:
                if attempt == max_retries - 1:
                    pass
                else:
                    await asyncio.sleep(2 * (attempt + 1))
            except Exception as e:
                break

        await asyncio.sleep(0.1)


# ---------------- Улучшенная проверка VK ----------------
async def check_vk_comments(context: CallbackContext):
    """Улучшенная функция проверки комментариев с обработкой ошибок"""
    global is_checking

    if is_checking:
        logger.info("🔁 Проверка уже выполняется, пропускаем...")
        return 0, 0

    is_checking = True
    found_count = 0
    processed_groups = 0
    total_checked_comments = 0
    total_checked_posts = 0

    try:
        groups = get_groups()
        keywords = get_keywords()

        if not groups:
            logger.warning("⚠️ Нет групп для проверки")
            return processed_groups, found_count

        if not keywords:
            logger.warning("⚠️ Нет ключевых слов для проверки")
            return processed_groups, found_count

        if not vk:
            logger.error("❌ VK API не инициализирован")
            return processed_groups, found_count

        logger.info(f"🔍 Начинаем проверку: {len(groups)} групп, {len(keywords)} ключевых слов")

        start_time = time.time()

        for domain, group_id in groups:
            try:
                processed_groups += 1
                group_comments_checked = 0
                group_comments_found = 0
                group_posts_checked = 0

                logger.info(f"📋 Проверяем группу: {domain} (ID: {group_id})")

                # Получаем посты со стены
                try:
                    posts = await safe_vk_request(
                        vk.wall.get,
                        owner_id=-group_id,
                        count=20,
                        filter="owner"
                    )
                    if not posts or 'items' not in posts:
                        logger.warning(f"  ⚠️ В группе {domain} нет постов или ошибка доступа")
                        continue

                    posts = posts['items']
                    group_posts_checked = len(posts)
                    total_checked_posts += len(posts)
                    logger.info(f"  📝 Получено {len(posts)} постов для проверки")

                    # Добавляем посты в Excel (без проверки на уникальность)
                    for post in posts:
                        post_text = post.get('text', '')
                        add_post_to_excel(domain, group_id, post['id'], post_text)

                except Exception as e:
                    logger.error(f"  ❌ Ошибка получения постов для {domain}: {e}")
                    continue

                for post in posts:
                    if post.get('comments', {}).get('count', 0) > 0:
                        try:
                            comments = await safe_vk_request(
                                vk.wall.getComments,
                                owner_id=-group_id,
                                post_id=post['id'],
                                count=100
                            )
                            if not comments or 'items' not in comments:
                                continue

                            comments = comments['items']
                            group_comments_checked += len(comments)
                            total_checked_comments += len(comments)

                        except Exception as e:
                            logger.warning(f"    ⚠️ Ошибка получения комментариев к посту {post['id']}: {e}")
                            continue

                        for comment in comments:
                            comment_id = comment.get('id')
                            if not comment_id:
                                continue

                            text = comment.get('text', '')
                            from_id = comment.get('from_id')

                            if from_id and from_id < 0:
                                continue

                            contains, found_keyword = contains_keyword(text, keywords)

                            if contains:
                                try:
                                    user_info = await safe_vk_request(
                                        vk.users.get,
                                        user_ids=from_id,
                                        fields="city,photo_200"
                                    )
                                    user_name = "Неизвестный пользователь"
                                    city = "не указан"
                                    photo_url = None

                                    if user_info:
                                        user_info = user_info[0]
                                        user_name = f"{user_info.get('first_name', '')} {user_info.get('last_name', '')}".strip()
                                        city = user_info.get("city", {}).get("title", "не указан")
                                        # Получаем URL аватарки
                                        photo_url = user_info.get('photo_200')

                                    group_link = f"https://vk.com/{domain}"
                                    post_link = f"https://vk.com/wall-{group_id}_{post['id']}?reply={comment_id}"
                                    user_link = f"https://vk.com/id{from_id}" if from_id else "не доступно"

                                    # Формируем текстовое сообщение с новым порядком полей
                                    text_message = (
                                        "⚡ Хром работал 24/7 и обнаружил комментарий, необходимо включиться!\n\n"
                                        f"💬 <b>Текст комментария:</b>\n"
                                        f"{user_name}: {text[:500]}\n\n"
                                        f"🔗 <b>Ссылка на страницу пользователя:</b> {user_link}\n"
                                        f"🌍 <b>Город:</b> {city}\n"
                                        f"🔗 <b>Ссылка на комментарий:</b> {post_link}\n"
                                        f"🔗 <b>Ссылка на группу:</b> {group_link}\n"
                                        f"🔍 <b>Маркер:</b> {found_keyword}"
                                    )

                                    # Подготавливаем данные для Excel в новом порядке
                                    comment_excel_data = {
                                        'user_name': user_name,
                                        'user_link': user_link,
                                        'city': city,
                                        'text': text,
                                        'comment_link': post_link,
                                        'keyword': found_keyword,
                                        'detection_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                    }

                                    # Добавляем комментарий в Excel
                                    add_comment_to_excel(comment_excel_data)

                                    await send_notification_with_photo(context, text_message, photo_url)
                                    increment_total_comments_count()
                                    found_count += 1
                                    group_comments_found += 1

                                    logger.info(f"    ✅ НАЙДЕН КОММЕНТАРИЙ: {user_name} - '{found_keyword}'")

                                except Exception as e:
                                    logger.error(f"    ❌ Ошибка обработки найденного комментария: {e}")

                # Логируем результаты по группе
                if group_comments_found > 0:
                    logger.info(
                        f"  ✅ Группа {domain}: проверено {group_posts_checked} постов, {group_comments_checked} комментариев, найдено {group_comments_found}")
                else:
                    logger.info(
                        f"  📊 Группа {domain}: проверено {group_posts_checked} постов, {group_comments_checked} комментариев, совпадений нет")

                await asyncio.sleep(0.5)

            except Exception as e:
                logger.error(f"❌ Критическая ошибка при проверке группы {domain}: {e}")
                continue

        # Итоговый отчет
        end_time = time.time()
        duration = end_time - start_time

        if found_count > 0:
            logger.info(
                f"🎉 ПРОВЕРКА ЗАВЕРШЕНА: обработано {processed_groups} групп, проверено {total_checked_posts} постов и {total_checked_comments} комментариев, найдено {found_count} совпадений за {duration:.1f} сек")
        else:
            logger.info(
                f"📊 ПРОВЕРКА ЗАВЕРШЕНА: обработано {processed_groups} групп, проверено {total_checked_posts} постов и {total_checked_comments} комментариев, совпадений не найдено за {duration:.1f} сек")

        return processed_groups, found_count

    except Exception as e:
        logger.error(f"💥 Критическая ошибка в функции проверки: {e}")
        return processed_groups, found_count
    finally:
        is_checking = False


# ---------------- Обработка сообщений ----------------
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Всегда разрешаем доступ
    user_input = update.message.text
    chat_id = update.message.chat_id
    chat_type = update.effective_chat.type
    message_text = user_input.lower()

    if message_text == "статус":
        groups = get_groups()
        keywords = get_keywords()
        chats = get_all_chats()

        # Проверяем статус текущего чата
        current_chat_status = "✅ добавлен" if is_chat_in_db(chat_id) else "❌ не добавлен"

        chats_info = []
        for cid, ctype, ctitle in chats:
            chat_desc = f"- {ctitle or 'Личный чат'} ({ctype}, ID: {cid})"
            chats_info.append(chat_desc)

        status_text = (
                f"📊 <b>Текущий статус:</b>\n\n"
                f"{get_bot_status()}\n\n"
                f"<b>Детальная информация:</b>\n"
                f"Группы ВК: {len(groups)}\n"
                f"Ключевые слова: {len(keywords)}\n"
                f"Чаты для уведомлений: {len(chats)}\n"
                f"Текущий чат: {current_chat_status}\n\n"
                + ("\n".join(chats_info) if chats_info else "Нет добавленных чатов.")
        )
        await update.message.reply_html(status_text, reply_markup=get_main_keyboard())

    elif message_text == "добавить группу":
        await update.message.reply_text(
            "Введите ссылку на группу ВКонтакте (например: https://vk.com/relaxmore1) или короткое имя:",
            reply_markup=get_main_keyboard())
        context.user_data['awaiting_input'] = 'group'

    elif message_text == "добавить ключевое слово":
        await update.message.reply_text("Введите ключевые слова через запятую:", reply_markup=get_main_keyboard())
        context.user_data['awaiting_input'] = 'keyword'

    elif message_text == "список групп":
        groups = get_groups()
        if groups:
            group_list = "\n".join([f"{i + 1}. {g[0]} (ID: {g[1]})" for i, g in enumerate(groups)])
            await update.message.reply_text(f"Отслеживаемые группы:\n{group_list}", reply_markup=get_main_keyboard())
        else:
            await update.message.reply_text("Список групп пуст.", reply_markup=get_main_keyboard())

    elif message_text == "список ключевых слов":
        keywords = get_keywords()
        await update.message.reply_text(
            "Ключевые слова:\n" + ("\n".join(keywords) if keywords else "Список ключевых слов пуст."),
            reply_markup=get_main_keyboard())

    elif message_text == "удалить группу":
        groups = get_groups()
        if groups:
            await update.message.reply_text(
                "Выберите группу для удаления:\n" + "\n".join([f"{i + 1}. {g[0]}" for i, g in enumerate(groups)]),
                reply_markup=get_main_keyboard())
            context.user_data['awaiting_input'] = 'delete_group'
        else:
            await update.message.reply_text("Список групп пуст.", reply_markup=get_main_keyboard())

    elif message_text == "удалить ключевое слово":
        keywords = get_keywords()
        if keywords:
            await update.message.reply_text("Выберите ключевое слово для удаления:\n" + "\n".join(
                [f"{i + 1}. {k}" for i, k in enumerate(keywords)]), reply_markup=get_main_keyboard())
            context.user_data['awaiting_input'] = 'delete_keyword'
        else:
            await update.message.reply_text("Список ключевых слов пуст.", reply_markup=get_main_keyboard())

    # НОВАЯ КНОПКА: Удалить все ключевые слова
    elif message_text == "удалить все ключевые слова":
        keywords = get_keywords()
        if keywords:
            delete_all_keywords()
            await update.message.reply_text(
                "✅ Все ключевые слова удалены!",
                reply_markup=get_main_keyboard()
            )
            logger.info("✅ Пользователь удалил все ключевые слова")
        else:
            await update.message.reply_text(
                "❌ Список ключевых слов и так пуст.",
                reply_markup=get_main_keyboard()
            )

    # НОВЫЕ КОМАНДЫ ДЛЯ УПРАВЛЕНИЯ ЧАТАМИ
    elif message_text == "добавить чат":
        current_chat_id = update.effective_chat.id
        chat_type = update.effective_chat.type
        chat_title = update.effective_chat.title

        if is_chat_in_db(current_chat_id):
            await update.message.reply_text(
                "✅ Этот чат уже добавлен для получения уведомлений!",
                reply_markup=get_main_keyboard()
            )
        else:
            add_chat_to_db(current_chat_id, chat_type, chat_title)
            await update.message.reply_text(
                "✅ Чат успешно добавлен для получения уведомлений!",
                reply_markup=get_main_keyboard()
            )
            logger.info(f"✅ Добавлен чат для уведомлений: {chat_title or 'Личный чат'} (ID: {current_chat_id})")

    elif message_text == "удалить чат":
        current_chat_id = update.effective_chat.id
        chat_title = update.effective_chat.title

        if is_chat_in_db(current_chat_id):
            remove_chat_from_db(current_chat_id)
            await update.message.reply_text(
                "✅ Чат удален из списка для уведомлений!",
                reply_markup=get_main_keyboard()
            )
            logger.info(f"❌ Удален чат из уведомлений: {chat_title or 'Личный чат'} (ID: {current_chat_id})")
        else:
            await update.message.reply_text(
                "❌ Этот чат не был добавлен для уведомлений.",
                reply_markup=get_main_keyboard()
            )

    elif message_text == "список чатов":
        chat_list_text = get_chats_list_text()
        await update.message.reply_text(chat_list_text, reply_markup=get_main_keyboard())

    elif message_text == "проверить сейчас":
        await update.message.reply_text("🔄 Запускаю проверку...", reply_markup=get_main_keyboard())
        logger.info("🔄 Ручная проверка запущена пользователем")
        processed_groups, found_count = await check_vk_comments(context)
        total_comments = get_total_comments_count()
        excel_posts, excel_comments = get_excel_stats()

        if found_count > 0:
            await update.message.reply_text(
                f"✅ Проверка завершена! Найдено {found_count} новых комментариев с ключевыми словами.\n"
                f"📈 Всего найдено: {total_comments}\n"
                f"📁 Постов в Excel: {excel_posts}\n"
                f"📁 Комментариев в Excel: {excel_comments}",
                reply_markup=get_main_keyboard())
        else:
            await update.message.reply_text(
                f"✅ Проверка завершена! Новых комментариев с ключевыми словами не найдено.\n"
                f"📈 Всего найдено: {total_comments}\n"
                f"📁 Постов в Excel: {excel_posts}\n"
                f"📁 Комментариев в Excel: {excel_comments}",
                reply_markup=get_main_keyboard())

    elif message_text == "экспорт в excel":
        """Новая команда для отправки Excel файлов"""
        try:
            excel_posts, excel_comments = get_excel_stats()

            # Форматируем файлы перед отправкой
            format_excel_file(POSTS_EXCEL_FILE)
            format_excel_file(COMMENTS_EXCEL_FILE)

            if excel_posts > 0 and os.path.exists(POSTS_EXCEL_FILE):
                with open(POSTS_EXCEL_FILE, 'rb') as posts_file:
                    await update.message.reply_document(
                        document=posts_file,
                        filename="checked_posts.xlsx",
                        caption=f"📊 Файл с проверенными постами\nКоличество записей: {excel_posts}"
                    )
            else:
                await update.message.reply_text("📭 Файл с постами пуст или не существует")

            if excel_comments > 0 and os.path.exists(COMMENTS_EXCEL_FILE):
                with open(COMMENTS_EXCEL_FILE, 'rb') as comments_file:
                    await update.message.reply_document(
                        document=comments_file,
                        filename="found_comments.xlsx",
                        caption=f"📊 Файл с найденными комментариями\nКоличество записей: {excel_comments}"
                    )
            else:
                await update.message.reply_text("📭 Файл с комментариями пуст или не существует")

        except Exception as e:
            await update.message.reply_text(f"❌ Ошибка при экспорте в Excel: {e}")

    elif 'awaiting_input' in context.user_data:
        input_type = context.user_data['awaiting_input']

        if input_type == 'group':
            groups = [g[0] for g in get_groups()]

            extracted_identifier = extract_group_id_from_url(user_input)

            if not extracted_identifier:
                await update.message.reply_text(
                    "❌ Не удалось извлечь идентификатор группы. Пожалуйста, введите корректную ссылку или имя группы.",
                    reply_markup=get_main_keyboard())
                context.user_data.pop('awaiting_input')
                return

            if extracted_identifier in groups:
                await update.message.reply_text("⚠️ Эта группа уже в списке!", reply_markup=get_main_keyboard())
            else:
                try:
                    group_info = await safe_vk_request(
                        vk.groups.getById,
                        group_id=extracted_identifier
                    )
                    if group_info:
                        group_info = group_info[0]
                        group_id = group_info['id']
                        add_group(extracted_identifier, group_id)
                        logger.info(f"✅ Добавлена группа: {extracted_identifier} (ID: {group_id})")
                        await update.message.reply_text(f"✅ Группа {extracted_identifier} (ID: {group_id}) добавлена!",
                                                        reply_markup=get_main_keyboard())
                    else:
                        await update.message.reply_text("❌ Не удалось получить информацию о группе",
                                                        reply_markup=get_main_keyboard())
                except vk_api.exceptions.ApiError as e:
                    if e.code == 100:
                        await update.message.reply_text("❌ Группа не найдена. Проверьте правильность ссылки.",
                                                        reply_markup=get_main_keyboard())
                    elif e.code == 15:
                        await update.message.reply_text("❌ Нет доступа к группе. Возможно, она приватная или удалена.",
                                                        reply_markup=get_main_keyboard())
                    else:
                        await update.message.reply_text(f"❌ Ошибка VK API: {e}", reply_markup=get_main_keyboard())
                except Exception as e:
                    await update.message.reply_text("❌ Ошибка добавления группы. Проверьте правильность ссылки.",
                                                    reply_markup=get_main_keyboard())
            context.user_data.pop('awaiting_input')

        elif input_type == 'keyword':
            keywords_input = user_input.split(',')
            added_count = 0
            existing_count = 0

            for kw in keywords_input:
                keyword = kw.strip()
                if keyword:
                    keywords = get_keywords()
                    if keyword not in keywords:
                        add_keyword(keyword)
                        added_count += 1
                        logger.info(f"✅ Добавлено ключевое слово: '{keyword}'")
                    else:
                        existing_count += 1

            if added_count > 0:
                await update.message.reply_text(f"✅ Добавлено {added_count} ключевых слов!",
                                                reply_markup=get_main_keyboard())
            if existing_count > 0:
                await update.message.reply_text(f"⚠️ {existing_count} слов уже были в списке!",
                                                reply_markup=get_main_keyboard())

            context.user_data.pop('awaiting_input')

        elif input_type == 'delete_group':
            groups = get_groups()
            try:
                index = int(user_input) - 1
                if 0 <= index < len(groups):
                    removed = groups[index][0]
                    delete_group(removed)
                    logger.info(f"❌ Удалена группа: {removed}")
                    await update.message.reply_text(f"❌ Группа {removed} удалена!", reply_markup=get_main_keyboard())
                else:
                    await update.message.reply_text("⚠️ Неверный номер группы!", reply_markup=get_main_keyboard())
            except ValueError:
                await update.message.reply_text("⚠️ Пожалуйста, введите номер группы!",
                                                reply_markup=get_main_keyboard())
            context.user_data.pop('awaiting_input')

        elif input_type == 'delete_keyword':
            keywords = get_keywords()
            try:
                index = int(user_input) - 1
                if 0 <= index < len(keywords):
                    removed = keywords[index]
                    delete_keyword(removed)
                    logger.info(f"❌ Удалено ключевое слово: '{removed}'")
                    await update.message.reply_text(f"❌ Ключевое слово '{removed}' удалено!",
                                                    reply_markup=get_main_keyboard())
                else:
                    await update.message.reply_text("⚠️ Неверный номер слова!", reply_markup=get_main_keyboard())
            except ValueError:
                await update.message.reply_text("⚠️ Пожалуйста, введите номер слова!", reply_markup=get_main_keyboard())
            context.user_data.pop('awaiting_input')
    else:
        if chat_type in ['group', 'supergroup']:
            await update.message.reply_text("Используйте кнопки для управления ботом",
                                            reply_markup=get_admin_keyboard())
        else:
            await update.message.reply_text("Используйте кнопки для управления ботом", reply_markup=get_main_keyboard())


# ---------------- Улучшенная периодическая проверка ----------------
async def periodic_check(context: CallbackContext):
    """Улучшенная функция периодической проверки с обработкой ошибок"""
    try:
        logger.info("⏰ Запуск автоматической проверки по расписанию")
        processed_groups, found_count = await check_vk_comments(context)

        # Итог автоматической проверки
        if found_count > 0:
            logger.info(f"🎯 АВТОПРОВЕРКА: найдено {found_count} новых комментариев")
        else:
            logger.info("📭 Автопроверка: новых комментариев не найдено")

    except Exception as e:
        logger.error(f"💥 Ошибка в автоматической проверке: {e}")


# ---------------- Проверка доступности VK API ----------------
def check_vk_api_availability():
    """Проверяет доступность VK API"""
    try:
        response = requests.get('https://api.vk.com/method/utils.getServerTime', timeout=10)
        return response.status_code == 200
    except:
        return False


# ---------------- Упрощенная функция запуска ----------------
def main():
    """Основная функция"""
    global bot_start_time

    # Проверка наличия токенов перед запуском
    if not VK_TOKEN:
        print("❌ ОШИБКА: VK_TOKEN не найден в переменных окружения!")
        print("   Создайте файл .env и укажите в нем VK_TOKEN=your_token")
        sys.exit(1)
    
    if not TELEGRAM_TOKEN:
        print("❌ ОШИБКА: TELEGRAM_TOKEN не найден в переменных окружения!")
        print("   Создайте файл .env и укажите в нем TELEGRAM_TOKEN=your_token")
        sys.exit(1)

    # Устанавливаем время запуска бота
    bot_start_time = datetime.now()

    # Инициализация базы данных
    init_db()

    # Инициализация Excel файлов
    init_excel_files()

    # Проверка доступности VK API
    if not check_vk_api_availability():
        print("✗ VK API недоступен")

    # Выводим сообщение о запуске бота
    print("=" * 50)
    print("🤖 БОТ ДЛЯ МОНИТОРИНГА VK КОММЕНТАРИЕВ")
    print("=" * 50)
    print(f"🚀 Запуск: {bot_start_time.strftime('%H:%M:%S')}")
    print(f"📊 Всего комментариев: {get_total_comments_count()}")
    print(f"📋 Групп ВК: {len(get_groups())}")
    print(f"🔍 Ключевых слов: {len(get_keywords())}")
    print(f"💬 Чатов для уведомлений: {len(get_all_chats())}")

    # Статистика Excel файлов
    excel_posts, excel_comments = get_excel_stats()
    print(f"📁 Постов в Excel: {excel_posts}")
    print(f"📁 Комментариев в Excel: {excel_comments}")

    print("⏰ Автопроверка каждые 10 минут")
    print("=" * 50)
    print("📝 Ожидание проверки...")
    print("=" * 50)

    try:
        # Создаем Application с включенным JobQueue
        application = Application.builder().token(TELEGRAM_TOKEN).build()

        # Хендлеры
        application.add_handler(CommandHandler("start", start))
        application.add_handler(CommandHandler("keyboard", keyboard_command))
        application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

        # Периодическая проверка каждые 10 минут
        job_queue = application.job_queue
        job_queue.run_repeating(
            periodic_check,
            interval=600,
            first=10,
            name="periodic_vk_check",
            job_kwargs={
                'misfire_grace_time': 300,
                'coalesce': True,
                'max_instances': 1
            }
        )

        # Запускаем бота с обработкой ошибок
        application.run_polling(
            poll_interval=1,
            timeout=30,
            drop_pending_updates=True
        )

    except NetworkError as e:
        print(f"Сетевая ошибка: {e}")

    except Exception as e:
        print(f"Ошибка: {e}")

    finally:
        print("Бот остановлен")


if __name__ == "__main__":
    main()